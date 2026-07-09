from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
ALT_FILL_1 = PatternFill("solid", fgColor="F8FBFF")
ALT_FILL_2 = PatternFill("solid", fgColor="EDF3FA")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def normalize_excel_value(val):
    if isinstance(val, list):
        return ", ".join(str(x) for x in val)
    if isinstance(val, tuple):
        return ", ".join(str(x) for x in val)
    if isinstance(val, set):
        return ", ".join(str(x) for x in val)
    if isinstance(val, dict):
        return str(val)
    if val is None:
        return ""
    return str(val)


def style_header_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_label_value(ws, row_idx, label_col=1, value_col=2):
    ws.cell(row=row_idx, column=label_col).font = Font(bold=True)
    ws.cell(row=row_idx, column=label_col).fill = SECTION_FILL
    ws.cell(row=row_idx, column=label_col).alignment = Alignment(vertical="center")
    ws.cell(row=row_idx, column=value_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row_idx, column=label_col).border = THIN_BORDER
    ws.cell(row=row_idx, column=value_col).border = THIN_BORDER


def autosize_columns(ws, min_width=12, max_width=40):
    dims = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            dims[cell.column] = max(dims.get(cell.column, min_width), min(len(value) + 2, max_width))

    for col_idx, width in dims.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_alt_rows(ws, start_row, end_row, end_col):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL_1 if r % 2 else ALT_FILL_2
        for c in range(1, end_col + 1):
            ws.cell(r, c).fill = fill
            ws.cell(r, c).border = THIN_BORDER


def build_excel(conversation, database: str = "") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conversations"

    ws["A1"] = "Database"
    ws["B1"] = database or "-"
    ws["A2"] = "Exported"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    style_label_value(ws, 1)
    style_label_value(ws, 2)

    header_row = 4
    headers = ["Nr", "Type", "Question", "Kind", "Intent", "Complexity / Notes"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)
    style_header_row(ws, header_row)

    row_idx = header_row + 1

    for idx, item in enumerate(conversation, start=1):
        classification = item.get("classification", {}) or {}
        answer = normalize_excel_value(item.get("answer", ""))
        sql_text = normalize_excel_value(item.get("sql", ""))

        notes = []
        if answer:
            notes.append(answer)
        if sql_text:
            notes.append("SQL:\n" + sql_text)

        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=normalize_excel_value(classification.get("type", classification.get("tip", ""))))
        ws.cell(row=row_idx, column=3, value=normalize_excel_value(item.get("question", "")))
        ws.cell(row=row_idx, column=4, value=normalize_excel_value(classification.get("kind", "")))
        ws.cell(row=row_idx, column=5, value=normalize_excel_value(classification.get("intent", "")))
        ws.cell(row=row_idx, column=6, value="\n\n".join(notes))

        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=c).border = THIN_BORDER

        row_idx += 1

    apply_alt_rows(ws, header_row + 1, row_idx - 1, 6)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 70

    for r in range(header_row + 1, row_idx):
        ws.row_dimensions[r].height = 52

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions

    for idx, item in enumerate(conversation, start=1):
        ws_chat = wb.create_sheet(title=f"Chat_{idx}")

        classification = item.get("classification", {}) or {}
        question = normalize_excel_value(item.get("question", ""))
        tip = normalize_excel_value(classification.get("type", classification.get("tip", "")))
        kind = normalize_excel_value(classification.get("kind", ""))
        intent = normalize_excel_value(classification.get("intent", ""))
        sql_text = normalize_excel_value(item.get("sql", ""))
        explanation = normalize_excel_value(item.get("answer", ""))

        meta_rows = [
            ("Question", question),
            ("Type", tip),
            ("Kind", kind),
            ("Intent", intent),
            ("SQL", sql_text),
            ("Explanation", explanation),
        ]

        current_row = 1
        for label, value in meta_rows:
            ws_chat.cell(row=current_row, column=1, value=label)
            ws_chat.cell(row=current_row, column=2, value=value)
            style_label_value(ws_chat, current_row)
            current_row += 1

        result = item.get("result")
        if isinstance(result, dict) and result.get("columns") and result.get("rows") is not None:
            columns = [normalize_excel_value(c) for c in result.get("columns", [])]
            data_rows = result.get("rows", [])

            current_row += 2
            table_header_row = current_row

            for col_idx, col_name in enumerate(columns, start=1):
                ws_chat.cell(row=table_header_row, column=col_idx, value=col_name)
            style_header_row(ws_chat, table_header_row)

            for r_idx, data_row in enumerate(data_rows[:50], start=table_header_row + 1):
                if isinstance(data_row, dict):
                    values = [normalize_excel_value(data_row.get(col, "")) for col in columns]
                else:
                    values = [normalize_excel_value(v) for v in data_row]

                for c_idx, value in enumerate(values, start=1):
                    ws_chat.cell(row=r_idx, column=c_idx, value=value)
                    ws_chat.cell(row=r_idx, column=c_idx).alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
                    ws_chat.cell(row=r_idx, column=c_idx).border = THIN_BORDER

            apply_alt_rows(ws_chat, table_header_row + 1, ws_chat.max_row, len(columns))
            ws_chat.freeze_panes = f"A{table_header_row + 1}"
            ws_chat.auto_filter.ref = ws_chat.dimensions

        ws_chat.column_dimensions["A"].width = 18
        ws_chat.column_dimensions["B"].width = 95
        autosize_columns(ws_chat, min_width=12, max_width=32)

        for r in range(1, ws_chat.max_row + 1):
            ws_chat.row_dimensions[r].height = 24

        for r in range(1, 7):
            ws_chat.row_dimensions[r].height = 38

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio