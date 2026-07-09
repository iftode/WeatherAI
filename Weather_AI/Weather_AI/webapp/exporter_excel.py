from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name: str) -> str:
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:31]


def build_excel(conversation, database: str = "") -> BytesIO:
    wb = Workbook()

    # Sheet 1 - overview
    ws = wb.active
    ws.title = "Conversations"

    ws["A1"] = "Database"
    ws["B1"] = database or ""

    ws["A2"] = "Exported"
    ws["B2"] = datetime.now().isoformat(timespec="seconds")

    ws["A4"] = "Nr"
    ws["B4"] = "Type"
    ws["C4"] = "Question"
    ws["D4"] = "Kind"
    ws["E4"] = "Intent"
    ws["F4"] = "Complexity / Notes"

    row_idx = 5

    for idx, item in enumerate(conversation, start=1):
        classification = item.get("classification", {})

        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=item.get("type", ""))
        ws.cell(row=row_idx, column=3, value=item.get("question", ""))
        ws.cell(row=row_idx, column=4, value=classification.get("kind", ""))
        ws.cell(row=row_idx, column=5, value=classification.get("intent", ""))

        if item.get("type") == "business":
            ws.cell(row=row_idx, column=6, value=item.get("answer", ""))
        else:
            ws.cell(row=row_idx, column=6, value=item.get("sql", ""))

        row_idx += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28

    # Separate sheet for each interaction
    for idx, item in enumerate(conversation, start=1):
        title = _safe_sheet_name(f"Chat_{idx}")
        sh = wb.create_sheet(title=title)

        sh["A1"] = "Question"
        sh["B1"] = item.get("question", "")

        classification = item.get("classification", {})
        sh["A2"] = "Type"
        sh["B2"] = item.get("type", "")
        sh["A3"] = "Kind"
        sh["B3"] = classification.get("kind", "")
        sh["A4"] = "Intent"
        sh["B4"] = classification.get("intent", "")

        if item.get("type") == "business":
            sh["A6"] = "Answer"
            sh["B6"] = item.get("answer", "")
            sh.column_dimensions["A"].width = 18
            sh.column_dimensions["B"].width = 100

        else:
            sh["A6"] = "SQL"
            sh["B6"] = item.get("sql", "")

            sh["A7"] = "Explanation"
            sh["B7"] = item.get("explanation", "")

            columns = item.get("columns", [])
            rows = item.get("rows", [])

            start_row = 9

            for c, name in enumerate(columns, start=1):
                sh.cell(row=start_row, column=c, value=name)

            for r_i, row in enumerate(rows, start=start_row + 1):
                for c_i, val in enumerate(row, start=1):
                    sh.cell(row=r_i, column=c_i, value=val)

            for c in range(1, max(2, len(columns)) + 1):
                sh.column_dimensions[get_column_letter(c)].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio